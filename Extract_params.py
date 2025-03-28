# Import the required function
import pandas as pd
import openpyxl
import csv

# Funtion to create a df from excel sheet
def ToDataFrame(excel_wb: str):
    # Read in the inventory sheet
    wb = openpyxl.load_workbook(excel_wb, 
                                data_only=True)
    # From excel to csv
    ws = wb['Farm Data - Grains']
    # Keep the record of the initial row
    initial_row = ws.max_row
    initial_col = ws.max_column
    # To store the headers
    Headers = []
    for row in range(1, initial_row + 1):
        Headers.append(ws.cell(row, 1).value)
    # To store the crop data
    for i in range(2, initial_col + 1):
        globals()[f'row{i}'] = []
        for row in range(1, initial_row + 1):
            globals()[f'row{i}'].append(ws.cell(row, i).value)
    # Write back into the excel sheet
    # To transform the header
    for col in range(1, initial_row + 1):
        Header = Headers[col - 1]
        ws.cell(1, col).value = Header
    # To transpose the data
    for i in range(2, initial_col + 1):
        rows = globals()[f'row{i}']
        for f, row in enumerate(rows):
            ws.cell(i, f + 1).value = row
    # Written into a csv
    with open('df.csv', 'w', newline="") as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])        
    # convert into dataframe object  
    df = pd.DataFrame(pd.read_csv('df.csv'))
    # Get rid of the extra row
    df = df.iloc[0:12]
    return df

# Separate the big df into crop type
def ByCropType(df: pd.DataFrame):
    Crop = []
    for i in df['Crop type'].index:
        Crop.append(df.iloc[i])
    return Crop

# Get the general information
def GenInfo(excel_wb: str):
    # Read in the excel wb
    wb = openpyxl.load_workbook(excel_wb, data_only=True)
    # Activate the sheet
    ws = wb['General information']
    # Get the state loc info (assuming the cell does not change)
    loc = ws.cell(9, 2).value
    # Rainfall > 600mm
    if ws.cell(17, 2).value == 'N':
        rain_over = False
    else:
        rain_over = True
    return loc, rain_over