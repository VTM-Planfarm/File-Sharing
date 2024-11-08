# Import the required pacakage
import openpyxl.cell
import openpyxl
import openpyxl.utils.dataframe
import openpyxl.cell.cell as Cell
import pandas as pd
from Extract_params import GenInfo, ToDataFrame, ByCropType
from From_q import FollowUp, ListFertChem, ToSoilAme, ToVeg, SpecCrop
import requests as rq
import streamlit as st
import shutil, os, tempfile
from datetime import datetime as dt

# Get current path
cwd = os.getcwd()

st.title("Carbon accounting")
st.header("Questionnaire extraction")

in_q = st.file_uploader("Upload your questionnaire form as a csv format:")

# Read in the form as csv
df = pd.read_csv(in_q)

# Number of crop in the questionnaire
crops = df['What crops did you grow last year?'].iloc[0].split('\n')

st.write(crops)

st.radio("Select your crop:", crops)

config = {}
config['filename'] = st.text_input("Name your file (client, location, etc.):")

if st.button("Do you want to check you questionnaire?"):
    st.dataframe(df)

if st.button("Run"):
    # Temp output folder
    tmp_out = tempfile.mkdtemp(prefix='output',dir=cwd)

    # Write out the general info
    FollowUp(df, tmp_out)

    # Crop specific info
    SpecCrop(df, crops)
    # Write into the inventory sheet
    wb = openpyxl.load_workbook("Inventory sheet v1 - Grain.xlsx")
    # Fill in general info
    ws = wb['General information']

    ## Business name & location & rf
    ws.cell(1, 2).value = df['Property name '].iloc[0]
    ws.cell(2, 2).value = df['Property location'].iloc[0]
    ws.cell(1, 11).value = df['Property average annual rainfall (mm)'].iloc[0]
    for i in range(2, 3):
        ws.cell(1, i + 1).value = df[f'Property {i} name '].iloc[0]
        ws.cell(2, i + 1).value = df[f'Property {i} location'].iloc[0]
        ws.cell(1, i + 10).value = df[f'Property {i} average annual rainfall (mm)'].iloc[0]

    ## Rainfall & request ETo from DPIRD
    # if df['Property average annual rainfall (mm)'].iloc[0] > 0:
    #     ws.cell(1, 11).value = df['Property average annual rainfall (mm)'].iloc[0]
    # else:
    #     # Request both rainfall and Eto from DPIRD
    #     pass

    CropType = Cell.Cell(ws, 9, 1)

    for i in range(12):
        CC = CropType.offset(i + 1)
        for crop in crops:
            if crop == CC.value:
                # Area sown
                CC.offset(column=1).value = df[f'What area was sown to {crop.lower()} last year? (Ha)'].iloc[0]
                # Last year yield
                CC.offset(column=2).value = df[f'What did your {crop.lower()} crop yield on average last year? (t/ha)'].iloc[0]
                # Fraction of crop burnt
                CC.offset(column=3).value = df[f'Was any land burned to prepare for {crop.lower()} crops last year? If so, how much? (Ha)'].iloc[0] / df[f'What area was sown to {crop.lower()} last year? (Ha)'].iloc[0]

    ## Electricity
    ws.cell(22, 5).value = df['Annual electricity usage last year (kwh)'].iloc[0]
    ws.cell(22, 6).value = float(df['Percentage of annual renewable electricity usage last year '].iloc[0].rstrip('%'))

    # Fertiliser
    ws = wb['Fertiliser Applied - Input']

    fert_applied = ListFertChem(df, crops, 1)

    for i, crop in enumerate(crops):
        fert = fert_applied[i]
        row = 2
        space = 0
        if i > 0:
            row += len(fert_applied[i-1].keys())
        for key, value in fert.items():
            # Product name
            ws.cell(row + space, 1).value = key
            # Rate
            ws.cell(row + space, 6).value = value[0]
            # Forms
            ws.cell(row + space, 2).value = value[1]
            # Crop
            ws.cell(row + space, 4).value = crop 
            space += 1

    # Chemical
    ws = wb['Chemical Applied - Input']

    chem_applied = ListFertChem(df, crops, 2)

    for i, crop in enumerate(crops):
        chem = chem_applied[i]
        row = 2
        space = 0
        if i > 0:
            row += len(chem_applied[i-1].keys())
        for key, value in chem.items():
            # Product name
            ws.cell(row + space, 1).value = key
            # Crop
            ws.cell(row + space, 15).value = crop
            # Rate
            ws.cell(row + space, 16).value = value[0]
            space += 1

    # Lime/gypsum
    ws = wb['Lime Product - Input']

    products_applied = ToSoilAme(df, crops)

    i = 0
    while i < len(products_applied[crop]) * len(crops):
        for crop in crops:
            for key, value in products_applied[crop].items():
                ws.cell(2 + i, 1).value = key
                ws.cell(2 + i, 3).value = crop
                ws.cell(2 + i, 5).value = value[0]
                ws.cell(2 + i, 4).value = value[1]
                i += 1

    #  Fuel usage
    ws = wb['Fuel Usage - Input']

    # Vegetation
    ws = wb['Vegetation - Input']

    vegetation = ToVeg(df)

    ws.cell(2, 2).value = vegetation['species']
    ws.cell(2, 3).value = vegetation['soil type']
    ws.cell(2, 4).value = vegetation['ha']
    ws.cell(2, 5).value = vegetation['age']

    wb.save(os.path.join(tmp_out, 'Inventory_Sheet.xlsx'))

    shutil.make_archive("Question_Extract", "zip", tmp_out)

    zip_name = config['filename'] + str(dt.today().strftime('%d-%m-%Y'))

    with open("Question_Extract.zip", "rb") as f:
        st.download_button('Download the extracted info', f, file_name=zip_name+".zip")

    os.remove('Question_Extract.zip')
    shutil.rmtree(tmp_out)

# General info
loc, rain_over, prod_sys = GenInfo('Inventory sheet v1 - Grain.xlsx')

# Create a df using function
df = ToDataFrame('Inventory sheet v1 - Grain.xlsx')

# Separate it by crop type
Crop = ByCropType(df)

# Selected crop
# 0: Wheat
# 1: Barley
# 2: Canola
# 3: Lupins
# 4: Oats
# 5: Hay
# 6: Triticale
# 7: Field Peas
# 8: Chick Peas
# 9: Faba Beans
# 10: Lentils
# 11: Other Grains
selected_crop = 0

# url and key
API_url = 'https://emissionscalculator-mtls.production.aiaapi.com/calculator/v1/grains'
# Add in the key and perm file when AIA gets back to us
cert = ('something.key', 'something.perm')

# Set the header
Headers = {
    'Authorisation': 'Bearer <token>',
    'Content-type': 'application/json',
    'User-Agent': 'Chrome/120.0.0.0',
    "Accept": "application/json"
}

# params for the API
datas = {
    'state': loc,
    'crops': [
        {
            'type': Crop[selected_crop]['Crop type'],
            'state': loc,
            'productionSystem': prod_sys,
            'averageGrainYield': Crop[selected_crop]['Average grain yield (t/ha)'],
            'areaSown': Crop[selected_crop]['Area sown (ha)'],
            'nonUreaNitrogen': Crop[selected_crop]['Non-Urea Nitrogen Applied (kg N/ha)'],
            'ureaApplication': Crop[selected_crop]['Urea Applied (kg Urea/ha)'],
            'ureaAmmoniumNitrate': Crop[selected_crop]['Urea-Ammonium Nitrate (UAN) (kg product/ha)'],
            'phosphorusApplication': Crop[selected_crop]['Phosphorus Applied (kg P/ha)'],
            'potassiumApplication': Crop[selected_crop]['Potassium Applied (kg K/ha)'],
            'sulfurApplication': Crop[selected_crop]['Sulfur Applied (kg S/ha)'],
            'rainfallAbove600': rain_over,
            'fractionOfAnnualCropBurnt': Crop[selected_crop]['Fraction of the annual production of crop that is burnt (%)'],
            'herbicideUse': Crop[selected_crop]['General Herbicide/Pesticide use (kg a.i. per crop)'],
            'glyphosateOtherHerbicideUse': Crop[selected_crop]['Herbicide (Paraquat, Diquat, Glyphoste) (kg a.i. per crop)'],
            'electricityAllocation': Crop[selected_crop]['electricityAllocation'],
            'limestone': Crop[selected_crop]['Mass of Lime Applied (total tonnes)'],
            'limestoneFraction': Crop[selected_crop]['Fraction of Lime/Dolomite'],
            'dieselUse': Crop[selected_crop]['Annual Diesel Consumption (litres/year)'],
            'petrolUse': Crop[selected_crop]['Annual Pertol Use (litres/year)'],
            'lpg': 0
        }
    ],
    'electricityRenewable': Crop[selected_crop]['% of electricity from renewable source'],
    'electricityUse': Crop[selected_crop]['Annual Electricity Use (state Grid) (KWh)'],
    'vegetation': [
        {
            'vegetation': {
                'region': Crop[selected_crop]['Region'],
                'treeSpecies': Crop[selected_crop]['Tree Species'],
                'soil': Crop[selected_crop]['Soil'],
                'area': Crop[selected_crop]['Area (ha)'],
                'age': Crop[selected_crop]['Age (yrs)']
            },
            'allocationToCrops': Crop[selected_crop]['Allocation to crop']
        }
    ]
}

# GET request for the API Grains only
# response = requests.post(url=API_url, 
#                         headers=Headers,
#                         data=datas,
#                         cert=cert)